"""
Excel export — extends the existing `write_workbook()` from
ai_risk_framework.py (all the raw-data and reconciliation tabs it already
builds) with three additional tabs this app adds on top: job-role-level
detail, pattern classification, and saved annotations. The resulting file
is saved directly to output/reports/ (see backend/storage.py) — it is not
a temp file, it's the persisted report the Export screen serves.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

import ai_risk_framework  # module import, not `from ... import SOC_REF` — see main.py
from ai_risk_framework import write_workbook, build_reconciliation
from backend.patterns import classify_pattern
from backend.job_roles import build_all_job_roles
from backend.annotations import load_annotations
from backend.view1_source_breakdown import build_view1
from backend.view2_comparison_matrix import build_view2

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")


def _write_df(wb, ws_name, df, col_widths=None, wrap_cols=None):
    ws = wb.create_sheet(ws_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            if wrap_cols and cell.column_letter in wrap_cols:
                cell.alignment = WRAP
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"
    return ws


def _write_risk_action_matrix_sheet(wb, comparison_df, matrix_df):
    """Write comparison_df and matrix_df as two clearly separated, labeled
    tables within one sheet — a bold section-header row above each."""
    ws = wb.create_sheet("Risk_Action_Matrix")

    ws.append(["Per-Occupation Comparison"])
    ws.cell(row=ws.max_row, column=1).font = Font(name="Arial", bold=True, size=12)
    header_row = ws.max_row + 1
    for col_idx, col in enumerate(comparison_df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in dataframe_to_rows(comparison_df, index=False, header=False):
        ws.append(row)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
            if cell.column_letter in {"E", "F", "G", "I"}:
                cell.alignment = WRAP

    ws.append([])
    ws.append(["Risk & Action Matrix (reference)"])
    ws.cell(row=ws.max_row, column=1).font = Font(name="Arial", bold=True, size=12)
    header_row2 = ws.max_row + 1
    for col_idx, col in enumerate(matrix_df.columns, start=1):
        cell = ws.cell(row=header_row2, column=col_idx, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in dataframe_to_rows(matrix_df, index=False, header=False):
        ws.append(row)
    for row in ws.iter_rows(min_row=header_row2 + 1, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
            if cell.column_letter == "E":
                cell.alignment = WRAP

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 34
    ws.column_dimensions["G"].width = 34
    ws.column_dimensions["I"].width = 60
    return ws


def _included_socs(annotations_df):
    """SOCs to include in the two app-added summary tabs, honoring the
    per-SOC 'Include this SOC in exported report' checkbox from Screen 3.
    A SOC with no saved annotation yet defaults to included (matches
    `annotations.get_annotation()`'s default)."""
    excluded = set(
        annotations_df.loc[~annotations_df["include_in_report"].astype(bool), "soc_code"]
    ) if not annotations_df.empty else set()
    return [soc for soc in ai_risk_framework.SOC_REF["soc_code"] if soc not in excluded]


def build_extended_workbook(path):
    """Write the full workbook: base tabs (from ai_risk_framework's own
    write_workbook, always the FULL 3-SOC dataset) plus Job_Roles_by_SOC,
    Pattern_Classification, and Annotations tabs. The two summary tabs
    respect each SOC's 'include in report' checkbox; the base tabs are the
    unmodified existing function's output and are not filtered."""

    # 1) Base workbook, unchanged logic from ai_risk_framework.py.
    write_workbook(path)

    # 2) Reopen and append this app's extra tabs.
    wb = load_workbook(path)

    annotations_df = load_annotations()
    included_socs = _included_socs(annotations_df)

    job_roles_df = build_all_job_roles(included_socs)
    if not job_roles_df.empty:
        _write_df(wb, "Job_Roles_by_SOC", job_roles_df, col_widths={"B": 30})

    reconciliation = build_reconciliation()
    pattern_rows = []
    for row in reconciliation.itertuples():
        if row.soc_code not in included_socs:
            continue
        pattern = classify_pattern(row._asdict())
        pattern_rows.append({
            "soc_code": row.soc_code,
            "soc_title": row.soc_title,
            "risk_rating": row.risk_rating,
            "pattern_name": pattern["pattern_name"],
            "meaning": pattern["meaning"],
            "recommended_action": pattern["recommended_action"],
            "assumptions": "; ".join(pattern["assumptions"]) if pattern["assumptions"] else "",
        })
    pattern_df = pd.DataFrame(pattern_rows)
    _write_df(wb, "Pattern_Classification", pattern_df,
              col_widths={"B": 22, "E": 60, "F": 50, "G": 50}, wrap_cols={"E", "F", "G"})

    if annotations_df.empty:
        annotations_df = pd.DataFrame(columns=["soc_code", "note", "include_in_report", "updated_at"])
    _write_df(wb, "Annotations", annotations_df, col_widths={"B": 60}, wrap_cols={"B"})

    # 3) Signal Ledger / Risk & Action Matrix tabs — additive only, computed
    # independently of everything above (they re-read input/*.csv
    # themselves; see backend/view1_source_breakdown.py and
    # backend/view2_comparison_matrix.py).
    signal_ledger_df = build_view1()
    _write_df(wb, "Signal_Ledger", signal_ledger_df,
              col_widths={"A": 22, "B": 24, "C": 12, "D": 14, "E": 70}, wrap_cols={"E"})

    comparison_df, matrix_df = build_view2()
    _write_risk_action_matrix_sheet(wb, comparison_df, matrix_df)

    wb.save(path)
    return path


def build_views_workbook(path):
    """A separate, lightweight workbook containing ONLY the Signal Ledger
    and Risk & Action Matrix sheets — for sharing when a user only needs
    these views, not the full reconciliation workbook. Same tab names/
    formatting as the tabs added to the consolidated workbook above;
    regenerated every pipeline run alongside it."""
    signal_ledger_df = build_view1()
    comparison_df, matrix_df = build_view2()

    wb = Workbook()
    wb.remove(wb.active)
    _write_df(wb, "Signal_Ledger", signal_ledger_df,
              col_widths={"A": 22, "B": 24, "C": 12, "D": 14, "E": 70}, wrap_cols={"E"})
    _write_risk_action_matrix_sheet(wb, comparison_df, matrix_df)

    wb.save(path)
    return path
