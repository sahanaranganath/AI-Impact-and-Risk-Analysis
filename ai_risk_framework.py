"""
AI Risk Framework — Data Build, Reconciliation & Insight Engine
=================================================================

Purpose
-------
Builds a per-SOC-code AI risk view for three occupations in the
SETT segment (Software Developers, Electrical Engineers, Chemical
Technicians) by combining:

  1. External labor market data   (BLS OEWS, BLS Employment Projections)
  2. AI impact data                (O*NET tasks, Anthropic Economic Index,
                                     OECD AI Capability Gap Index)
  3. Internal company data         (ATS/VMS reqs, client billing,
                                     candidate pipeline — DUMMY DATA)

...then reconciles all sources into one table per SOC code and
generates rule-based insights / recommendations, before writing
everything to a single Excel workbook with multiple tabs.

IMPORTANT — DATA PROVENANCE NOTES
----------------------------------
- External data values below were manually transcribed from screenshots
  the user shared of BLS, O*NET, Anthropic, and OECD sources. Where a
  cell text was ambiguous in a screenshot, the closest legible reading
  was used. These are NOT live API pulls — treat as a snapshot.
- Internal data (ATS/VMS, billing, pipeline) is 100% SYNTHETIC/DUMMY
  data created for demonstration of the reconciliation logic. It does
  NOT represent any real company's actual data.
- OECD does not cover SOC 15-1252 (Software Developers) directly in
  its published dataset. As a documented workaround, this script uses
  15-1251.00 (Computer Programmers) — the nearest adjacent occupation
  in the SOC hierarchy — as a PROXY, and flags it clearly wherever it
  is used. This is an estimation, not real OECD data for that SOC.
"""

import os
import pandas as pd
from datetime import datetime

pd.set_option("display.width", 140)

# =====================================================================
# All data tables below are loaded from CSV files under input/, not
# hardcoded in this script. This is the single place that reads them —
# add a new SOC by adding matching rows to these CSVs (soc_ref.csv plus
# the corresponding external/internal rows) and every downstream
# reconciliation, agent, dashboard, and report picks it up automatically,
# with no code changes required.
# =====================================================================

INPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "input")


def _read_csv(name, **kwargs):
    return pd.read_csv(os.path.join(INPUT_DIR, f"{name}.csv"), **kwargs)


def load_all_data():
    """(Re)read every input/*.csv file into the module-level DataFrames
    below. Called once at import time, and again at the start of every
    `build_reconciliation()` call — so editing the CSVs (e.g. adding a
    new SOC) is picked up by the next pipeline run with no server
    restart and no code changes required."""
    global SOC_REF, BLS_OEWS, BLS_PROJECTIONS, ONET_OCCUPATION, ONET_TASKS
    global ANTHROPIC_JOB_EXPOSURE, ANTHROPIC_TASK_PENETRATION, OECD_CAPABILITY_GAP
    global ATS_VMS_REQS, CLIENT_BILLING, CANDIDATE_PIPELINE
    global EXTERNAL_SOURCES_REF, INTERNAL_SOURCES_REF

    # ---- SECTION 0 — SOC reference (master key every table joins on) ----
    SOC_REF = _read_csv("soc_ref")
    # O*NET-style code (SOC + ".00") used by O*NET / Anthropic / OECD files —
    # derived, not stored, so adding a SOC only requires soc_code + soc_title.
    SOC_REF["onet_code"] = SOC_REF["soc_code"] + ".00"

    # ---- SECTION 1 — EXTERNAL DATA ----
    # 1a. BLS OEWS (Occupational Employment & Wage Statistics), May 2025.
    #     Source: bls.gov/oes — values transcribed from user screenshot.
    BLS_OEWS = _read_csv("bls_oews")

    # 1b. BLS Employment Projections, 2024-34 (Table 1.2), thousands.
    #     Source: bls.gov/emp/tables.htm
    BLS_PROJECTIONS = _read_csv("bls_projections")

    # 1c. O*NET Occupation Data (title + description).
    #     Source: onetcenter.org — transcribed from user screenshot.
    ONET_OCCUPATION = _read_csv("onet_occupation")

    # 1d. O*NET Task Statements — captured in full for Software Developers
    #     only (15-1252.00); other SOCs are not represented in this file.
    #     This does not block SOC-level reconciliation (which uses
    #     job-level exposure scores), only the task-level drill-down.
    ONET_TASKS = _read_csv("onet_tasks")

    # 1e. Anthropic Economic Index — Job Exposure (occupation-level).
    #     Usage-based, observed exposure. 0-1 scale, higher = more exposed.
    #     Source: huggingface.co/datasets/Anthropic/EconomicIndex
    ANTHROPIC_JOB_EXPOSURE = _read_csv("anthropic_job_exposure")

    # 1f. Anthropic Economic Index — Task Penetration, matched to
    #     Software Developers' O*NET tasks only (same coverage limit as
    #     ONET_TASKS above).
    ANTHROPIC_TASK_PENETRATION = _read_csv("anthropic_task_penetration")

    # 1g. OECD AI Capability Gap Index (Data sheet, baseline scenario).
    #     Reversed/normalized score: 0-1, higher = more exposed.
    #     NOTE: OECD's public dataset does not include 15-1252 directly;
    #     15-1251.00 (Computer Programmers) is used as a PROXY, flagged
    #     in the `is_proxy` column.
    OECD_CAPABILITY_GAP = _read_csv("oecd_capability_gap")

    # ---- SECTION 2 — INTERNAL DATA (100% SYNTHETIC / DUMMY) ----
    # 2a. ATS/VMS Req Data — job requisitions, fill status, timing.
    #     Source system this simulates: Applicant Tracking / Vendor
    #     Management System. Tells you DEMAND volume & fill difficulty.
    ATS_VMS_REQS = _read_csv("ats_vms_reqs", parse_dates=["open_date", "fill_date"])

    # 2b. Client Billing Data — bill rate / pay rate / margin per req.
    #     Source system this simulates: Finance / billing system.
    #     Tells you the PRICE signal (compression = possible AI pressure
    #     or commoditization; expansion = scarcity / premium skill).
    CLIENT_BILLING = _read_csv("client_billing")

    # 2c. Candidate Pipeline Data — submittals, interviews, offers.
    #     Source system this simulates: Recruiting CRM.
    #     Tells you the SUPPLY signal (thin pipeline = scarcity or
    #     shrinking interest in the role).
    CANDIDATE_PIPELINE = _read_csv("candidate_pipeline")

    # 2d. Internal / External source reference tables (Name + Description)
    #     — mirrors the "Sources" sheet style the user built manually.
    EXTERNAL_SOURCES_REF = _read_csv("external_sources_ref")
    INTERNAL_SOURCES_REF = _read_csv("internal_sources_ref")


load_all_data()


# =====================================================================
# SECTION 3 — RECONCILIATION LOGIC
# =====================================================================

def build_internal_summary():
    """Aggregate the three internal (dummy) tables to one row per SOC."""
    reqs = ATS_VMS_REQS.copy()
    billing = CLIENT_BILLING.copy()
    pipeline = CANDIDATE_PIPELINE.copy()

    rows = []
    for soc in SOC_REF["soc_code"]:
        soc_reqs = reqs[reqs["soc_code"] == soc].sort_values("open_date")
        soc_billing = billing[billing["soc_code"] == soc]
        soc_pipeline = pipeline[pipeline["soc_code"] == soc]

        total_reqs = len(soc_reqs)
        open_reqs = (soc_reqs["fill_status"] == "Open").sum()
        avg_days_to_fill = soc_reqs["days_to_fill"].mean(skipna=True)

        # Simple trend proxy: split reqs chronologically into first half vs
        # second half by open_date, compare req volume and avg margin.
        half = max(1, total_reqs // 2)
        early = soc_reqs.iloc[:half]
        late = soc_reqs.iloc[half:]

        early_ids = early["req_id"]
        late_ids = late["req_id"]
        early_margin = billing[billing["req_id"].isin(early_ids)]["margin"].mean()
        late_margin = billing[billing["req_id"].isin(late_ids)]["margin"].mean()
        margin_trend_pct = ((late_margin - early_margin) / early_margin * 100
                             if early_margin else 0)

        early_submittals = pipeline[pipeline["req_id"].isin(early_ids)]["submittals"].mean()
        late_submittals = pipeline[pipeline["req_id"].isin(late_ids)]["submittals"].mean()
        submittal_trend_pct = ((late_submittals - early_submittals) / early_submittals * 100
                                if early_submittals else 0)

        demand_trend_pct = ((len(late) - len(early)) / max(len(early), 1)) * 100

        rows.append({
            "soc_code": soc,
            "internal_total_reqs": total_reqs,
            "internal_open_reqs": int(open_reqs),
            "internal_avg_days_to_fill": round(avg_days_to_fill, 1),
            "internal_avg_bill_rate": round(soc_billing["bill_rate"].mean(), 1),
            "internal_avg_pay_rate": round(soc_billing["pay_rate"].mean(), 1),
            "internal_avg_margin": round(soc_billing["margin"].mean(), 1),
            "internal_margin_trend_pct": round(margin_trend_pct, 1),
            "internal_avg_submittals": round(soc_pipeline["submittals"].mean(), 1),
            "internal_submittal_trend_pct": round(submittal_trend_pct, 1),
            "internal_req_volume_trend_pct": round(demand_trend_pct, 1),
        })
    return pd.DataFrame(rows)


def compute_composite_ai_exposure(anthropic_df, oecd_df):
    """Merge Anthropic's usage-based exposure with OECD's capability-gap
    index on soc_code and average the two into one composite AI exposure
    score per SOC. Pulled out as a shared helper so `build_reconciliation()`
    below and `backend/view2_comparison_matrix.py` compute this number
    identically instead of each duplicating the formula."""
    merged = anthropic_df[["soc_code", "observed_exposure"]].merge(
        oecd_df[["soc_code", "gap_index_reversed_norm"]], on="soc_code", how="left"
    )
    merged["composite_ai_exposure"] = (
        merged["observed_exposure"] + merged["gap_index_reversed_norm"]
    ) / 2
    return merged[["soc_code", "composite_ai_exposure"]]


# =====================================================================
# SECTION 3B — SHARED BUSINESS-LOGIC THRESHOLDS
# =====================================================================
# Single source of truth for every LOW/MODERATE/HIGH exposure band and
# UP/DOWN direction call used across `classify_risk()` below,
# `backend/patterns.py`, `backend/agents.py`'s AI Impact Agent, and
# `backend/view2_comparison_matrix.py`'s Risk & Action Matrix — so "HIGH"
# (or "UP") means the same thing everywhere in the app, instead of several
# independently-reasoned cutoffs quietly disagreeing with each other.
#
# --- Composite AI exposure (0-1 scale) ---
# `composite_ai_exposure` is the mean of Anthropic's real, observed-usage
# exposure and OECD's forward-looking capability-gap score. These two
# inputs have very different natural ranges: Anthropic's usage-based
# figure is the differentiating "is this actually happening today" signal
# and rarely exceeds ~0.30-0.35 for any occupation in the published
# Economic Index, while OECD's capability-gap score asks "COULD AI
# eventually do this" and runs high (roughly 0.7-0.95) for most
# professional/technical occupations — it barely differentiates between
# them. Averaged together, the realistic range for a knowledge-work
# occupation sits roughly between ~0.30 (low real usage + moderate
# capability) and ~0.65 (elevated real usage + high capability) — not the
# full 0-1 range a generic threshold would assume. A cutoff as low as 0.25
# (the previous default) is at the FLOOR of that range, so it read almost
# every occupation as "HIGH" purely off OECD's near-constant ceiling,
# even when Anthropic's real-usage number was small. The bands below are
# calibrated to the composite's actual achievable range instead.
AI_EXPOSURE_LOW_MAX = 0.40    # below this: AI is a background factor, not yet a primary driver
AI_EXPOSURE_HIGH_MIN = 0.60   # at/above this: both real usage and forward capability agree — a primary risk driver
# 0.40-0.60: MODERATE — the two sources point in different directions or
# neither is strongly conclusive on its own.

# --- External demand direction (BLS 10-yr projected employment change %) ---
# National employment across ALL occupations grows roughly 3-4% per
# decade on average; BLS itself treats ~8%+ as "faster than average" job
# growth. A role projected at +2-3% isn't meaningfully "expanding" from a
# business standpoint — it's tracking the economy-wide baseline, not
# outperforming it.
WEAK_GROWTH_THRESHOLD = 8     # below this: not clearing the "faster than average" bar
STRONG_GROWTH_THRESHOLD = 10  # at/above this: a genuinely strong growth story

# --- Internal demand trend bands (early-vs-late req volume / margin trend %) ---
# The internal metric is a small-sample, noisier signal than BLS's
# national projection, so a small single-digit wobble shouldn't flip the
# read from "growing" to "declining" — a double-digit swing is needed
# before it's treated as a real trend rather than noise.
INTERNAL_TREND_FLAT_BAND = 10.0     # within +/-10%: flat/noise, not a clear trend
INTERNAL_SHARP_DECLINE_MAX = -10.0  # at/below this: sharply declining (used by patterns.py's structural-displacement check)


def classify_risk(ai_exposure, bls_growth_pct, internal_req_trend_pct, internal_margin_trend_pct):
    """
    Rule-based composite risk classification.

    Logic:
      - HIGH risk    : AI exposure high (>=AI_EXPOSURE_HIGH_MIN) AND
                        internal demand or margin trending down AND
                        external BLS growth is low/flat
                        (<WEAK_GROWTH_THRESHOLD).
      - LOW risk     : AI exposure low (<AI_EXPOSURE_LOW_MAX) OR (external
                        growth strong (>=STRONG_GROWTH_THRESHOLD) AND
                        internal demand trending up).
      - MODERATE     : everything else (mixed signals).
    """
    high_exposure = ai_exposure >= AI_EXPOSURE_HIGH_MIN
    low_exposure = ai_exposure < AI_EXPOSURE_LOW_MAX
    weak_growth = bls_growth_pct < WEAK_GROWTH_THRESHOLD
    strong_growth = bls_growth_pct >= STRONG_GROWTH_THRESHOLD
    internal_declining = internal_req_trend_pct < 0 or internal_margin_trend_pct < 0
    internal_growing = internal_req_trend_pct > 0 and internal_margin_trend_pct >= 0

    if high_exposure and weak_growth and internal_declining:
        return "HIGH"
    if low_exposure or (strong_growth and internal_growing):
        return "LOW"
    return "MODERATE"


def generate_insight(row):
    """Produce a plain-English narrative for one SOC row of the
    reconciled table, covering demand, price, supply, and AI exposure
    dimensions, plus a recommendation."""

    soc = row["soc_title"]
    parts = []

    # AI exposure framing
    parts.append(
        f"AI exposure signals: Anthropic observed usage-based exposure is "
        f"{row['anthropic_observed_exposure']:.2f} (0-1 scale), and OECD's "
        f"capability-gap index (higher = more exposed) is "
        f"{row['oecd_gap_index_reversed_norm']:.2f}"
        + (" [PROXY: OECD score uses 15-1251 Computer Programmers, not a "
           "direct 15-1252 measurement]" if row["oecd_is_proxy"] else "")
        + f". Composite AI exposure score: {row['composite_ai_exposure']:.2f}."
    )

    # External labor market framing
    growth_dir = "growing" if row["bls_emp_change_pct"] > 0 else "flat or shrinking"
    parts.append(
        f"External labor market: BLS projects national employment {growth_dir} "
        f"at {row['bls_emp_change_pct']:.1f}% over 2024-34, with "
        f"{row['bls_annual_openings_k']:.1f}k annual openings and a median wage "
        f"of ${row['bls_median_wage_2024']:,.0f}."
    )

    # Internal demand framing
    req_dir = ("increasing" if row["internal_req_volume_trend_pct"] > 0
               else "decreasing" if row["internal_req_volume_trend_pct"] < 0
               else "flat")
    parts.append(
        f"Internal demand: req volume is {req_dir} "
        f"({row['internal_req_volume_trend_pct']:+.1f}% early-vs-late period), "
        f"average time-to-fill is {row['internal_avg_days_to_fill']:.0f} days, "
        f"and {row['internal_open_reqs']} of {row['internal_total_reqs']} reqs "
        f"remain open."
    )

    # Internal price/margin framing
    margin_dir = ("expanding" if row["internal_margin_trend_pct"] > 0
                  else "compressing" if row["internal_margin_trend_pct"] < 0
                  else "stable")
    parts.append(
        f"Internal price signal: average margin is {margin_dir} "
        f"({row['internal_margin_trend_pct']:+.1f}%), average bill rate "
        f"${row['internal_avg_bill_rate']:.0f}/hr vs. pay rate "
        f"${row['internal_avg_pay_rate']:.0f}/hr."
    )

    # Internal supply framing
    supply_dir = ("growing" if row["internal_submittal_trend_pct"] > 0
                  else "shrinking" if row["internal_submittal_trend_pct"] < 0
                  else "stable")
    parts.append(
        f"Candidate supply: submittal volume per req is {supply_dir} "
        f"({row['internal_submittal_trend_pct']:+.1f}%)."
    )

    # Cross-source agreement check
    ext_signals_high_risk = row["bls_emp_change_pct"] < 8
    int_signals_declining = (row["internal_req_volume_trend_pct"] < 0
                              or row["internal_margin_trend_pct"] < 0)
    if row["risk_rating"] == "HIGH":
        agreement = ("External and internal signals AGREE this SOC warrants "
                      "attention: AI exposure is elevated, national growth is "
                      "modest, and internal demand/pricing shows early signs "
                      "of softening.")
    elif row["risk_rating"] == "LOW":
        agreement = ("Signals are broadly reassuring: either AI exposure is low, "
                      "or strong external growth is matched by healthy internal "
                      "demand and pricing.")
    else:
        agreement = ("Signals are MIXED — some indicators point toward risk "
                      "while others do not. Treat as a watch-list item rather "
                      "than an immediate action item.")
    parts.append(agreement)

    # Recommendation
    if row["risk_rating"] == "HIGH":
        rec = (f"RECOMMENDATION: Proactively flag {soc} to account managers and "
               f"talent teams. Consider (a) upskilling current bench toward "
               f"adjacent higher-complexity tasks within this role, (b) "
               f"reviewing bill-rate strategy before clients push for reductions "
               f"citing AI productivity gains, and (c) monitoring req volume "
               f"monthly rather than quarterly.")
    elif row["risk_rating"] == "LOW":
        rec = (f"RECOMMENDATION: Continue standard investment in {soc}. No "
               f"immediate AI-driven action needed; revisit at the next "
               f"quarterly review cycle.")
    else:
        rec = (f"RECOMMENDATION: Add {soc} to a quarterly watch-list. Re-run "
               f"this reconciliation in 3-6 months to see whether internal "
               f"demand/pricing trends firm up in either direction.")
    parts.append(rec)

    return " ".join(parts)


def build_reconciliation():
    """Join all external + internal sources into one row per SOC and
    compute the composite risk rating + narrative insight.

    Re-reads input/*.csv first, so a SOC added to those files since the
    last call (or since the server started) is picked up immediately —
    no restart needed."""
    load_all_data()

    df = SOC_REF.merge(BLS_OEWS.drop(columns=["occ_title"]), on="soc_code", how="left")
    df = df.merge(BLS_PROJECTIONS.drop(columns=["occ_title"]), on="soc_code", how="left")
    df = df.merge(
        ANTHROPIC_JOB_EXPOSURE.drop(columns=["title"]), on="soc_code", how="left"
    )
    df = df.merge(
        OECD_CAPABILITY_GAP[["soc_code", "gap_index_reversed_norm", "is_proxy"]]
        .rename(columns={"gap_index_reversed_norm": "oecd_gap_index_reversed_norm",
                          "is_proxy": "oecd_is_proxy"}),
        on="soc_code", how="left"
    )

    internal_summary = build_internal_summary()
    df = df.merge(internal_summary, on="soc_code", how="left")

    composite = compute_composite_ai_exposure(ANTHROPIC_JOB_EXPOSURE, OECD_CAPABILITY_GAP)
    df = df.merge(composite, on="soc_code", how="left")

    df["risk_rating"] = df.apply(
        lambda r: classify_risk(
            r["composite_ai_exposure"], r["emp_change_pct"],
            r["internal_req_volume_trend_pct"], r["internal_margin_trend_pct"]
        ), axis=1
    )

    df = df.rename(columns={
        "occ_title": "bls_occ_title_x",
        "observed_exposure": "anthropic_observed_exposure",
        "emp_change_pct": "bls_emp_change_pct",
        "annual_openings_k": "bls_annual_openings_k",
        "median_wage_2024": "bls_median_wage_2024",
    })

    df["insight_narrative"] = df.apply(generate_insight, axis=1)
    return df


# =====================================================================
# SECTION 4 — EXCEL OUTPUT
# =====================================================================

def write_workbook(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    reconciliation = build_reconciliation()
    internal_summary = build_internal_summary()

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    def write_df(ws_name, df, col_widths=None, wrap_cols=None):
        ws = wb.create_sheet(ws_name)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                if wrap_cols and cell.column_letter in wrap_cols:
                    cell.alignment = wrap
        if col_widths:
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width
        ws.freeze_panes = "A2"
        return ws

    # --- Reference / source description tabs ---
    write_df("External_Sources_Ref", EXTERNAL_SOURCES_REF,
             col_widths={"A": 32, "B": 90}, wrap_cols={"B"})
    write_df("Internal_Sources_Ref", INTERNAL_SOURCES_REF,
             col_widths={"A": 28, "B": 90}, wrap_cols={"B"})

    # --- Raw external data tabs ---
    write_df("BLS_OEWS", BLS_OEWS, col_widths={"B": 24})
    write_df("BLS_Projections", BLS_PROJECTIONS, col_widths={"B": 24, "H": 22})
    write_df("ONET_Occupation", ONET_OCCUPATION,
             col_widths={"B": 24, "C": 80}, wrap_cols={"C"})
    write_df("ONET_Tasks_SoftwareDev", ONET_TASKS,
             col_widths={"B": 20, "C": 70}, wrap_cols={"C"})
    write_df("Anthropic_Job_Exposure", ANTHROPIC_JOB_EXPOSURE, col_widths={"B": 24})
    write_df("Anthropic_Task_Pen_SWDev", ANTHROPIC_TASK_PENETRATION)
    write_df("OECD_Capability_Gap", OECD_CAPABILITY_GAP,
              col_widths={"C": 28})

    # --- Raw internal (dummy) data tabs ---
    write_df("ATS_VMS_Reqs", ATS_VMS_REQS, col_widths={"B": 28, "D": 14, "K": 14})
    write_df("Client_Billing", CLIENT_BILLING)
    write_df("Candidate_Pipeline", CANDIDATE_PIPELINE)
    write_df("Internal_Summary_by_SOC", internal_summary, col_widths={"A": 12})

    # --- Reconciliation tab ---
    recon_cols = [
        "soc_code", "soc_title", "onet_code",
        "tot_emp", "h_mean", "a_mean",
        "emp_2024_k", "emp_2034_k", "bls_emp_change_pct", "bls_annual_openings_k",
        "bls_median_wage_2024", "typical_education",
        "anthropic_observed_exposure", "oecd_gap_index_reversed_norm", "oecd_is_proxy",
        "composite_ai_exposure",
        "internal_total_reqs", "internal_open_reqs", "internal_avg_days_to_fill",
        "internal_avg_bill_rate", "internal_avg_pay_rate", "internal_avg_margin",
        "internal_margin_trend_pct", "internal_avg_submittals",
        "internal_submittal_trend_pct", "internal_req_volume_trend_pct",
        "risk_rating",
    ]
    recon_display = reconciliation[recon_cols].copy()
    ws = write_df("Reconciliation", recon_display, col_widths={"B": 22})
    # Color the risk_rating column for quick visual scan
    risk_col_idx = recon_cols.index("risk_rating") + 1
    risk_colors = {"HIGH": "F4CCCC", "MODERATE": "FFF2CC", "LOW": "D9EAD3"}
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=risk_col_idx)
        fill_color = risk_colors.get(cell.value)
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color,
                                     fill_type="solid")

    # --- Insights & Recommendations tab ---
    insights_df = reconciliation[["soc_code", "soc_title", "risk_rating",
                                   "insight_narrative"]].rename(
        columns={"insight_narrative": "insight_and_recommendation"}
    )
    write_df("Insights_Recommendations", insights_df,
             col_widths={"B": 22, "C": 14, "D": 120}, wrap_cols={"D"})
    ws_ins = wb["Insights_Recommendations"]
    for r in range(2, ws_ins.max_row + 1):
        ws_ins.row_dimensions[r].height = 220

    wb.save(path)
    return path


if __name__ == "__main__":
    OUT_PATH = "/mnt/user-data/outputs/AI_Risk_Framework_SETT.xlsx"
    write_workbook(OUT_PATH)
    print(f"Workbook written to: {OUT_PATH}")

    # Quick console preview of the reconciliation table
    recon = build_reconciliation()
    print("\n=== RECONCILIATION SUMMARY ===")
    print(recon[["soc_code", "soc_title", "composite_ai_exposure",
                  "bls_emp_change_pct", "internal_req_volume_trend_pct",
                  "internal_margin_trend_pct", "risk_rating"]].to_string(index=False))
